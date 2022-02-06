from py import process
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font
import pyodbc 
import os
import pandas as pd
import datetime
import tabula
from pathlib import Path
import shutil
import sys
import configparser


def processPDF(filepath):
    print("Processing PDF...")
    tabula.convert_into(filepath, v['rfq_folder'] + '/' + rfq_number + ".csv", output_format="csv", pages='all')
    read_file = pd.read_csv(v['rfq_folder'] + '/' + rfq_number + ".csv")
    read_file.to_excel (v['rfq_folder'] + '/' + rfq_number + '_temp.xlsx', index = None, header=True)
    
    wb = load_workbook(v['rfq_folder'] + '/' + rfq_number + '_temp.xlsx')
    ws = wb.worksheets[0]
    records = []
    row_no = 1
    for row in ws.rows:
        row_items = []
        for col in row:
            row_items.append(col.value)
        records.append(row_items)
        if len(records) > 1:
            if records[0] == row_items:
                ws.delete_rows(row_no, 1)
        row_no+=1
    wb.save(v['rfq_folder'] + '/' + rfq_number + '_to_process.xlsx')
    processExcel(v['rfq_folder'] + '/' + rfq_number + '_to_process.xlsx')

    files_to_remove = [
        v['rfq_folder'] + '/' + rfq_number + ".csv",
        v['rfq_folder'] + '/' + rfq_number + '_temp.xlsx',
        v['rfq_folder'] + '/' + rfq_number + '_to_process.xlsx'
    ]
    for file in files_to_remove:
        os.remove(file)

def processExcel(filepath):
    print("Processing Excel")
    wb = load_workbook(filepath)
    ws = wb.worksheets[0]
                    
    start_row = 0
    level_col_index = None
    part_no_col_index = None
    description_col_index = None
    qty_col_index = None
    uom_col_index = None

    level_col_list = v['rfq_to_db_mapping']['level'].split(',')
    level_col_list = [selected_column.lower().strip() for selected_column in level_col_list] 


    part_no_col_list = v['rfq_to_db_mapping']['part_no'].split(',')
    part_no_col_list = [selected_column.lower().strip() for selected_column in part_no_col_list] 

    description_col_list = v['rfq_to_db_mapping']['description'].split(',')
    description_col_list = [selected_column.lower().strip() for selected_column in description_col_list] 

    qty_list = v['rfq_to_db_mapping']['qty'].split(',')
    qty_list = [selected_column.lower().strip() for selected_column in qty_list] 
    
    uom_col_list = v['rfq_to_db_mapping']['uom'].split(',')
    uom_col_list = [selected_column.lower().strip() for selected_column in uom_col_list] 

    all_col_list = level_col_list + part_no_col_list + description_col_list + qty_list + uom_col_list

    start_row_list = []
    current_row_no = 1
    for row in ws.rows:
        for col in row:
            if str(col.value).lower().strip() in all_col_list:
                start_row_list.append(current_row_no)
        current_row_no+=1
    
    if len(start_row_list) != 0:
        check_if_quotation_exist = cursor.execute("select * from dbo.quotation where quotation_no=?", rfq_number)
        if len(check_if_quotation_exist.fetchall()) == 0:
            cursor.execute("insert into dbo.quotation(quotation_no, customer_email, assigned_staff, rfq_date, status) values (?, ?, ?, ?, ?)", rfq_number, v['customer_email'], 2, datetime.datetime.now(), 'draft')
            conn.commit()
            start_row = max(set(start_row_list), key = start_row_list.count)    
            
            all_column_headers_in_excel = []
            for col_name in ws[start_row]:
                all_column_headers_in_excel.append(str(col_name.value).lower().strip())
            
            for level_col in level_col_list:
                try:
                    level_col_index = all_column_headers_in_excel.index(level_col)
                    break
                except ValueError as e:
                    pass               
            for part_no_col in part_no_col_list:
                try:
                    part_no_col_index = all_column_headers_in_excel.index(part_no_col)
                    break
                except ValueError as e:
                    pass
            for description_col in description_col_list:
                try:
                    description_col_index = all_column_headers_in_excel.index(description_col)
                    break
                except ValueError as e:
                    pass
            for qty_col in qty_list:
                try:
                    qty_col_index = all_column_headers_in_excel.index(qty_col)
                    break
                except ValueError as e:
                    pass
            for uom_col in uom_col_list:
                try:
                    uom_col_index = all_column_headers_in_excel.index(uom_col)
                    break
                except ValueError as e:
                    pass
            
            current_row_no = 1
            row_unique_key_no = 0

            for row in ws.rows:
                if start_row < current_row_no:
                    level = row[level_col_index].value
                    part_no = row[part_no_col_index].value
                    description = row[description_col_index].value
                    qty = row[qty_col_index].value
                    uom = row[uom_col_index].value
                    is_bom = 0 

                    row_unique_key_no+=1

                    try:
                        current_level = float(level)
                    except:
                        if str(row[level_col_index].value).startswith('.'):
                            current_level = float(str(row[level_col_index].value).replace('.', ''))
                    
                    next_level = None
                    try:
                        next_level = float(ws[current_row_no+1][level_col_index].value)
                    except:
                        if str(ws[current_row_no+1][level_col_index].value).startswith('.'):
                            next_level = float(str(ws[current_row_no+1][level_col_index].value).replace('.', ''))

                    is_bom = 0
                    if row_unique_key_no == 1:
                        if next_level != None:
                            if next_level > current_level:
                                is_bom = 1
                        ## INITIAL INSERT
                        cursor.execute("insert into dbo.quotation_component(row, quotation_no, component_no, lvl, uom, description, quantity, unit_price, is_bom, bom_id) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", 
                                            row_unique_key_no, rfq_number, part_no, current_level, uom, description, qty, None, is_bom, None)
                        conn.commit()
                    else:
                        ## AFTER THAT, WE KEEP TRACKING BACK AND COMPARE THE CURRENT LEVEL TO ALL THE PREVIOUS LEVEL. THE NEAREST PREVIOUS LEVEL WITH A LEVEL LOWER THAN CURRENT LEVEL IS OUR SET! < ONLY APPLIES TO NON LEVEL 1
                        if next_level != None:
                            if next_level > current_level:
                                is_bom = 1
                        
                        found = False
                        curr_row = row_unique_key_no-1
                        while curr_row > 0:
                            row_level = cursor.execute("select id, lvl, component_no from dbo.quotation_component where row=? and quotation_no=?", curr_row, rfq_number)
                            result = row_level.fetchone()
                            if result != None:
                                if current_level > result.lvl:
                                    cursor.execute("insert into dbo.quotation_component(row, quotation_no, component_no, lvl, uom, description, quantity, unit_price, is_bom, bom_id) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", 
                                            row_unique_key_no, rfq_number, part_no, current_level, uom, description, qty, None, is_bom, result.id)
                                    conn.commit()
                                    found = True
                                    break
                            curr_row-=1
                        if found == False:
                            cursor.execute("insert into dbo.quotation_component(row, quotation_no, component_no, lvl, uom, description, quantity, unit_price, is_bom, bom_id) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", 
                                    row_unique_key_no, rfq_number, part_no, current_level, uom, description, qty, None, is_bom, None)
                            conn.commit()

                current_row_no+=1     
    else:
        print("None of the columns you specified are found")

config = configparser.ConfigParser()
config.read('sql_connect.cfg')

conn = pyodbc.connect('Driver=' + config['database']['driver'] + ';'
                      'Server=' + config['database']['server'] + ';'
                      'Database=' + config['database']['database'] + ';'
                      'Trusted_Connection=' + config['database']['trusted_connection'] + ';')
cursor = conn.cursor()

with open('aqs_bot/parts_extraction/config.yaml') as f:
    data = yaml.load(f, Loader=yaml.FullLoader)
    for key, v in data.items():
        customerlist = cursor.execute("select * from dbo.customer where company_email=?", v['customer_email'])
        if len(customerlist.fetchall()) == 0:
            cursor.execute("insert into dbo.customer(company_email, company_name, company_website) values (?, ?, ?)", v['customer_email'], v['customer'], v['customer_website'])
            conn.commit()
        ## Scan Makino folder for RFQs and process each of them
        rfq_list = os.listdir(v['rfq_folder'])
        for each_rfq in rfq_list:
            rfq_location = v['rfq_folder'] + '/' + each_rfq
            to_archive_rfq_location = v['archive_folder'] + '/' + each_rfq
            to_move_unprocessed_rfq_location = v['unprocessed_folder'] + '/' + each_rfq

            rfq_number = Path(rfq_location).stem
            rfq_file_extension = os.path.splitext(rfq_location)[1]
            
            if rfq_file_extension == ".pdf":
                try:
                    processPDF(rfq_location)
                    shutil.move(rfq_location, to_archive_rfq_location)
                except Exception as e:
                    print(e)
                    shutil.move(rfq_location, to_move_unprocessed_rfq_location)
            elif rfq_file_extension == '.xlsx':
                try:
                    processExcel(rfq_location)
                    shutil.move(rfq_location, to_archive_rfq_location)
                except Exception as e:
                    print(e)
                    shutil.move(rfq_location, to_move_unprocessed_rfq_location)
            else:
                shutil.move(rfq_location, to_move_unprocessed_rfq_location)
            
conn.close()
print("Connection closed.")