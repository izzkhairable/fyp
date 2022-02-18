from py import process
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

import pyodbc 
import os
import pandas as pd
import datetime
import tabula
from pathlib import Path
import shutil
import sys
import json 
import configparser

config = configparser.ConfigParser()
config.read('sql_connect.cfg')

conn = pyodbc.connect('Driver=' + config['database']['driver'] + ';'
                      'Server=' + config['database']['server'] + ';'
                      'Database=' + config['database']['database'] + ';'
                      'Trusted_Connection=' + config['database']['trusted_connection'] + ';')
cursor = conn.cursor()
wb = Workbook()

ws = wb.active
quotation = cursor.execute(
    "select labour_cost, markup_pct, labour_no_of_hours, testing_cost, remark from dbo.quotation where quotation_no=?", 'quotation_one')
res = quotation.fetchone()

markup_pct = None
if res.markup_pct != None:
    markup_pct = res.markup_pct/100
with open('aqs_bot/quotation_generation/config.yaml') as f:
    data = yaml.load(f, Loader=yaml.FullLoader)
    level_index = None 
    row_index = None
    part_no_index = None
    object_description_index = None
    quantity_index = None
    uom_index = None
    unit_price_index = None
    total_price_index = None
    remarks_index = None

    index = 0
    columns = []
    for i in list(data['column_configuration'].items()):
        if i[0] == "Level":
            level_index = index
        if i[0] == "Row":
            row_index = index
        if i[0] == "Part_No":
            part_no_index = index
        if i[0] == "Object_Description":
            object_description_index = index
        if i[0] == "Quantity":
            quantity_index = index
        if i[0] == "UOM":
            uom_index = index
        if i[0] == "Unit_Price":
            unit_price_index = index
        if i[0] == "Total_Price":
            total_price_index = index
        if i[0] == "Remarks":
            remarks_index = index
        columns.append(i[1])
        index+=1

    thin = Side(border_style="thin", color="000000")

    for i in range(1, len(columns)+1):
        ws.cell(row=1, column=i).value = columns[i-1]
        ws.cell(row=1, column=i).fill = PatternFill(start_color='4BACC6', end_color='4BACC6', fill_type="solid")
        ws.cell(row=1, column=i).font = Font(bold=True)
        ws.cell(row=1, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)


    row_level = cursor.execute(
        "select row, lvl, component_no, description, quantity, uom, unit_price, remark, crawl_info from dbo.quotation_component where quotation_no=? ORDER BY row ASC", 'quotation_one')
    result = row_level.fetchall()


    row_no = 2
    index_no = 1
    for each_row in result:
        multiple_supplier = False
        qty_location = get_column_letter(quantity_index+1) + str(row_no)
        unit_price_location = get_column_letter(unit_price_index+1) + str(row_no)
        
            
        if each_row.crawl_info != None:
            crawl_info = json.loads(each_row.crawl_info)
            if len(crawl_info) > 1:
                for ea_supplier in crawl_info:
                    qty_location = get_column_letter(quantity_index+1) + str(row_no)
                    unit_price_location = get_column_letter(unit_price_index+1) + str(row_no)
                    for i in range(1, len(columns)+1):
                        ws.cell(row=row_no, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.cell(row=row_no, column=level_index+1).value = each_row.lvl
                    ws.cell(row=row_no, column=row_index+1).value = index_no
                    ws.cell(row=row_no, column=part_no_index+1).value = each_row.component_no
                    ws.cell(row=row_no, column=object_description_index+1).value = each_row.description
                    ws.cell(row=row_no, column=quantity_index+1).value = ea_supplier['qty']
                    ws.cell(row=row_no, column=uom_index+1).value = each_row.uom
                    ws.cell(row=row_no, column=unit_price_index+1).value = (ea_supplier['unit_price'] * markup_pct) + ea_supplier['unit_price']
                    ws.cell(row=row_no, column=total_price_index+1).value = '=SUM(' + qty_location + "*" + unit_price_location + ')'
                    ws.cell(row=row_no, column=remarks_index+1).value = ea_supplier['supplier']

                    row_no+=1
                    index_no+=1
                multiple_supplier = True

        if multiple_supplier == False:
            
            ws.cell(row=row_no, column=level_index+1).value = each_row.lvl
            ws.cell(row=row_no, column=row_index+1).value = index_no
            ws.cell(row=row_no, column=part_no_index+1).value = each_row.component_no
            ws.cell(row=row_no, column=object_description_index+1).value = each_row.description
            ws.cell(row=row_no, column=quantity_index+1).value = each_row.quantity
            ws.cell(row=row_no, column=uom_index+1).value = each_row.uom
            ws.cell(row=row_no, column=unit_price_index+1).value = (each_row.unit_price * markup_pct) + each_row.unit_price
            ws.cell(row=row_no, column=total_price_index+1).value = '=SUM(' + qty_location + "*" + unit_price_location + ')'
            ws.cell(row=row_no, column=remarks_index+1).value = each_row.remark

            for i in range(1, len(columns)+1):
                ws.cell(row=row_no, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)

            row_no+=1
            index_no+=1

    start_total_price_location = get_column_letter(total_price_index+1) + str(2)
    end_total_price_location = get_column_letter(total_price_index+1) + str(row_no-1)


    row_no+=1
    qty_location = get_column_letter(quantity_index+1) + str(row_no)
    unit_price_location = get_column_letter(unit_price_index+1) + str(row_no)

    ws.cell(row=row_no, column=part_no_index+1).value = "LABOUR COST"
    ws.cell(row=row_no, column=object_description_index+1).value = "ASSEMBLY"
    ws.cell(row=row_no, column=quantity_index+1).value = res.labour_no_of_hours
    ws.cell(row=row_no, column=uom_index+1).value = "HR"
    ws.cell(row=row_no, column=unit_price_index+1).value = res.labour_cost
    ws.cell(row=row_no, column=total_price_index+1).value = "=" + qty_location + "*" + unit_price_location
    labour_cost_location = get_column_letter(total_price_index+1) + str(row_no)
    # ws.cell(row=row_no, column=9).value = res.remark
    row_no+=1
    ws.cell(row=row_no, column=unit_price_index+1).value = "TESTING COST"
    ws.cell(row=row_no, column=total_price_index+1).value = res.testing_cost
    testing_cost_location = get_column_letter(total_price_index+1) + str(row_no)

    row_no+=1
    ws.cell(row=row_no, column=unit_price_index+1).value = "SELLING PRICE"
    ws.cell(row=row_no, column=total_price_index+1).value = "=SUM(" + start_total_price_location + ":" + end_total_price_location + ")+" + labour_cost_location + "+" + testing_cost_location

    wb.save('aqs_bot/quotation_generation/document.xlsx')