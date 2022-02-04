from py import process
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

import pyodbc 
import os
import pandas as pd
import datetime
import tabula
from pathlib import Path
import shutil
import sys
import json 

conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=DESKTOP-KMU57HS;'
                      'Database=myerp101;'
                      'Trusted_Connection=yes;')
cursor = conn.cursor()
wb = Workbook()

ws = wb.active
quotation = cursor.execute(
    "select labour_cost, markup_pct, labour_cost_description from dbo.quotation where quotation_no=?", 'quotation_one')
res = quotation.fetchone()

markup_pct = None
if res.markup_pct != None:
    markup_pct = res.markup_pct/100

columns = ['Row', 'Component Number', 'Object Description', 'Qty', 'Component unit', 'Price', 'Total', 'Remarks']

for i in range(1, len(columns)+1):
    ws.cell(row=1, column=i).value = columns[i-1]


thin = Side(border_style="thin", color="000000")

for i in range(1,9):
    ws.cell(row=1, column=i).fill = PatternFill(start_color='4BACC6', end_color='4BACC6', fill_type="solid")
    ws.cell(row=1, column=i).font = Font(bold=True)
    ws.cell(row=1, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)

row_level = cursor.execute(
    "select row, component_no, description, quantity, uom, unit_price, remark, crawl_info from dbo.quotation_component where quotation_no=?", 'quotation_one')
result = row_level.fetchall()


row_no = 2
index_no = 1
for each_row in result:
    multiple_supplier = False
    qty_location = get_column_letter(4) + str(row_no)
    unit_price_location = get_column_letter(6) + str(row_no)
    
        
    if each_row.crawl_info != None:
        crawl_info = json.loads(each_row.crawl_info)
        if len(crawl_info) > 1:
            for ea_supplier in crawl_info:
                row_values = [
                index_no, 
                each_row.component_no, 
                each_row.description, 
                ea_supplier['qty'], 
                each_row.uom, 
                (ea_supplier['unit_price'] * markup_pct) + ea_supplier['unit_price'], 
                '=SUM(' + qty_location + "*" + unit_price_location + ')', 
                ea_supplier['supplier']
                ]

                for i in range(1, len(row_values)+1):
                    ws.cell(row=row_no, column=i).value = row_values[i-1]
                    ws.cell(row=row_no, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)

                row_no+=1
                index_no+=1
            multiple_supplier = True

    if multiple_supplier == False:
        row_values = [
        index_no, 
        each_row.component_no, 
        each_row.description, 
        each_row.quantity, 
        each_row.uom, 
        (each_row.unit_price * markup_pct) + each_row.unit_price, 
        '=SUM(' + qty_location + "*" + unit_price_location + ')', 
        each_row.remark
        ]

        for i in range(1, len(row_values)+1):
            ws.cell(row=row_no, column=i).value = row_values[i-1]
            ws.cell(row=row_no, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)

        row_no+=1
        index_no+=1

start_total_price_location = get_column_letter(7) + str(2)
end_total_price_location = get_column_letter(7) + str(row_no-1)


row_no+=1
ws.cell(row=row_no, column=6).value = "LABOUR COST"
ws.cell(row=row_no, column=7).value = res.labour_cost
labour_cost_location = get_column_letter(7) + str(row_no)
ws.cell(row=row_no, column=8).value = res.labour_cost_description
row_no+=2
ws.cell(row=row_no, column=6).value = "SELLING PRICE"
ws.cell(row=row_no, column=7).value = "=SUM(" + start_total_price_location + ":" + end_total_price_location + ")+" + labour_cost_location

wb.save('document.xlsx')